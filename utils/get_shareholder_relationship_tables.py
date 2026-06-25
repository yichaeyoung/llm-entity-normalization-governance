'''
DART 전자공시시스템에서 각 기업에 대한 주주관계테이블을 불러옵니다

<입력>
사업보고서 바로가기 링크

<동작 사항> 
DART 내부에서 주주에 관한 사항 테이블 추출 후 가족 관계 열을 추가하여 현재 프로젝트 내부의

"./datas/shareholder_relationship_tables/"

부분에

"기업명_주주관계.csv"

형태로 저장합니다

'''

import os 
import re
import time
import html
import requests
import pandas as pd
from openpyxl import load_workbook
from io import StringIO
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from tqdm import tqdm
from typing import Optional, List
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager


'''
INPUT_DATA_PATH : 사업보고서 바로가기 링크를 가져올 수 있는 파일
INPUT_EXCEL_PATH : 친족 지분 수식을 얻을 수 있는 파일이 있는 경우
OUTPUT_DIR_PATH : 테이블 저장 경로 (root directory)
'''
# [0] Path Setting
INPUT_DATA_PATH = '../datas/자기자본(DART).csv'
INPUT_EXCEL_PATH = '../datas/자기자본(DART).xlsx'
OUTPUT_DIR_PATH = '../datas/shareholder_relationship_tables'
OUTPUT_DIR_PATH_CURRENT = '../datas/shareholder_relationship_tables_current'


'''
HEADLESS : 서버 환경이면 True 로 설정
TARGET_TEXT : 찾고자 하는 테이블 명
TARGET_TABLE_SUFFIX : 저장할 테이블 식별자
WAIT_AFTER_VIEWER : viewer 페이지 로딩 대기 시간
FAMILY_REL_TOL : 허용 오차 범위 설정
'''
# [0] Others
HEADLESS = True
TARGET_TEXT = "VII. 주주에 관한 사항"
TARGET_TABLE_SUFFIX = "table_002"
WAIT_AFTER_VIEWER = 0.8
FAMILY_REL_TOL = 0.003  # 0.3%


# [1] 디렉토리 생성 부
os.makedirs(OUTPUT_DIR_PATH, exist_ok=True)


# [2] 함수 정의

def maybe_flatten_three_header_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    현재처럼:
        컬럼: 1번째 헤더줄 (성 명, 관 계, 주식의 종류, 소유주식수 및 지분율, ...)
        0번째 row: 2번째 헤더줄 (성 명, 관 계, 주식의 종류, 기 초, 기 초, 기 말, ...)
        1번째 row: 3번째 헤더줄 (성 명, 관 계, 주식의 종류, 주식수, 지분율, 주식수, ...)

    인 형태라면, 이 3줄을 합쳐서 컬럼 이름으로 만들고,
    위 2줄(0,1 row)을 날린 뒤 나머지를 데이터로 사용하는 함수.
    """
    # row가 너무 적으면 처리 X
    if df.shape[0] < 2:
        return df

    cols = [str(c).strip() for c in df.columns]
    row0 = [str(x).strip() for x in df.iloc[0].tolist()]
    row1 = [str(x).strip() for x in df.iloc[1].tolist()]

    # 3줄 헤더 패턴인지 간단히 검사 (앞 3컬럼 정도만 비교)
    is_three_header_like = True
    for idx in range(min(3, len(cols))):
        if not (cols[idx] == row0[idx] == row1[idx]):
            is_three_header_like = False
            break

    if not is_three_header_like:
        # 평범한 테이블이면 그대로 반환
        return df

    # ---- 여기부터는 3줄 헤더라고 가정하고 flatten ----
    header0 = cols
    header1 = row0
    header2 = row1

    new_cols = []
    for a, b, c in zip(header0, header1, header2):
        parts = [str(a).strip(), str(b).strip(), str(c).strip()]
        parts = [
            p for p in parts
            if p and p.lower() != "nan" and p != "-"
        ]
        if not parts:
            new_cols.append("")
        else:
            new_cols.append("_".join(parts))

    # 위의 두 줄(0,1)은 헤더였으니 날리고, 2번째 줄부터 데이터로 사용
    df_out = df.iloc[2:].copy()
    df_out.columns = new_cols
    df_out.reset_index(drop=True, inplace=True)
    return df_out

    
'''
오차 범위 제어 함수
'''
def is_close_to_any_family_number(
    x: int,
    family_numbers: list[int],
    rel_tol: float = FAMILY_REL_TOL,
    abs_tol: int = 5,  # 최소 허용 절대 오차 5주 예시
) -> bool:
    if not family_numbers:
        return False

    for n in family_numbers:
        if n == 0:
            if x == 0:
                return True
            continue

        diff = abs(x - n)
        if diff <= abs(n) * rel_tol or diff <= abs_tol:
            return True

    return False

'''
url 정보 정리 및 개선
'''
def normalize_viewer_url(url: str) -> str:
    if not url:
        return url
    u = html.unescape(url)
    u = re.sub(r'([?&])amp;', r'\1', u, flags=re.IGNORECASE)
    return u.strip()


'''
"VII. 주주에 관한 사항" 테이블 스트립트 블록 탐색
'''
def find_tree_block_for_target(soup: BeautifulSoup, target_text: str) -> Optional[str]:
    scripts = soup.find_all("script")
    for script in scripts:
        txt = script.string
        if not txt:
            continue
        if target_text in txt:
            return txt
    return None

'''
JS 블록 내 특정 제목 노드 추출
'''
def extract_viewer_params_from_tree_block(block: str, target_text: str) -> Optional[dict]:
    pattern_node = re.compile(
        r"node\d+\['text'\]\s*=\s*\"" + re.escape(target_text) + r"\";.*?treeData\.push\(node\d+\);",
        re.DOTALL
    )
    m = pattern_node.search(block)
    if not m:
        return None
    node_block = m.group(0)
    keys = ["rcpNo", "dcmNo", "eleId", "offset", "length", "dtd"]
    extracted = {}
    for k in keys:
        pm = re.search(rf"\['{k}'\]\s*=\s*\"([^\"]+)\"", node_block)
        if pm:
            extracted[k] = pm.group(1)
        else:
            return None
    return extracted


'''
DART viewer.do 접속용 URL 생성
'''
def build_viewer_url(params: dict) -> str:
    return (
        f"https://dart.fss.or.kr/report/viewer.do?"
        f"rcpNo={params['rcpNo']}&"
        f"dcmNo={params['dcmNo']}&"
        f"eleId={params['eleId']}&"
        f"offset={params['offset']}&"
        f"length={params['length']}&"
        f"dtd={params['dtd']}"
    )

'''
Selenium Chrome Brower Driver 생성
'''
def build_chrome_driver(headless=True):
    opts = ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1200,900")
    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    return driver


'''
Table 탐색 및 추출
'''
def find_html_with_tables_anywhere(driver, max_depth=5) -> str:
    def dfs(current_depth) -> Optional[str]:
        try:
            driver.find_element(By.CSS_SELECTOR, "table")
            time.sleep(0.12)
            return driver.page_source
        except Exception:
            pass

        if current_depth <= 0:
            return None
        frames = driver.find_elements(By.CSS_SELECTOR, "iframe")
        for f in frames:
            try:
                driver.switch_to.frame(f)
                got = dfs(current_depth - 1)
                if got:
                    return got
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    pass
        return None

    driver.switch_to.default_content()
    result = dfs(max_depth)
    return result or driver.page_source

'''
HTML 내 테이블 선택 및 저장
'''
def extract_and_save_only_target_table(html_text: str, out_dir: str, company: str, file_prefix: str = "", target_suffix: str = "table_002") -> int:
    soup = BeautifulSoup(html_text, "lxml")
    for br in soup.find_all("br"):
        br.replace_with("\n")
    tables = soup.find_all("table")
    if not tables:
        return 0

    os.makedirs(out_dir, exist_ok=True)
    saved = 0

    
    safe_company = re.sub(r"[^\w가-힣]", "_", company)

    for i, tbl in enumerate(tables, start=1):
        save_this = False
        tid = tbl.get("id", "")
        if tid:
            if re.search(rf"{re.escape(target_suffix)}$", tid):
                save_this = True
        else:
            m = re.match(r"table_(\d+)$", target_suffix)
            if m:
                target_idx = int(m.group(1))
                if i == target_idx:
                    save_this = True

        if not save_this:
            continue

        try:
            dfs = pd.read_html(StringIO(str(tbl)), flavor="lxml")
            if not dfs:
                continue
            df = dfs[0]
            df = df.apply(lambda s: s.map(lambda x: str(x).replace("\r", "").strip() if pd.notnull(x) else x))
            path = os.path.join(out_dir, f"{safe_company}_{file_prefix}{target_suffix}.csv")
            df.to_csv(path, index=False, encoding="utf-8-sig")
            print(f"saved: {path}")
            saved += 1
        except Exception as e:
            print(f"table#{i} 저장 실패: {e}")
            continue
    return saved


def get_family_share_numbers(excel_path: str, company: str, formula_col_idx: int = 14):
    """
    엑셀 파일에서 해당 회사의 '친족지분(주식수)' 수식/값에서 숫자만 추출하여 리스트로 반환.
    - 수식 형태: =1271876+547071+...
    - 단일 상수: =9507590   또는   9507590   또는   "9507590"
    """
    wb = load_workbook(excel_path, data_only=False)
    ws = wb.active

    target_name = str(company).strip()

    for row in ws.iter_rows(min_row=2):
        c = row[0].value  # 기업명
        if not c:
            continue
        if str(c).strip() != target_name:
            continue

        formula_cell = row[formula_col_idx].value

        # 1) 값이 아예 없으면 실패
        if formula_cell is None:
            return []

        # 2) 숫자 타입인 경우: 단일 상수로 간주
        if isinstance(formula_cell, (int, float)):
            try:
                return [int(formula_cell)]
            except Exception:
                return []

        # 3) 문자열인 경우: 맨 앞의 '='는 떼고, 안 떼도 되고 상관 없이 숫자 추출
        if isinstance(formula_cell, str):
            s = formula_cell.strip()
            # "=12345+6789" → "12345+6789"
            if s.startswith("="):
                s = s[1:]

            nums = re.findall(r"\d+", s)
            if not nums:
                return []

            return [int(n) for n in nums]

        # 4) 그 외 타입은 지원 안 함
        return []

    # 회사 이름을 못 찾은 경우
    return []


'''
HTML 내 테이블 선택 및 저장 + 친족여부 열 추가 저장
'''
def extract_and_save_only_target_table_family(
    html_text: str,
    out_dir: str,
    company: str,
    excel_path: str,
    file_prefix: str = "",
    target_suffix: str = "table_002",
) -> int:
    """
    1) HTML에서 target_suffix에 해당하는 table만 추출
    2) DataFrame으로 변환 후 컬럼 정리
    3) 엑셀(xlsx)의 '친족지분(주식수)' 수식을 기반으로 '친족여부' 열 생성 ('친족' / '아님')
    4) 최종 CSV 저장

    반환값: 저장된 테이블 개수 (일반적으로 0 또는 1)
    """
    soup = BeautifulSoup(html_text, "lxml")

    # <br> 태그를 줄바꿈으로 치환
    for br in soup.find_all("br"):
        br.replace_with("\n")

    tables = soup.find_all("table")
    if not tables:
        return 0

    os.makedirs(out_dir, exist_ok=True)
    saved = 0

    # 기업명 안전하게 변환
    safe_company = re.sub(r"[^\w가-힣]", "_", company)

    # 엑셀에서 해당 회사의 친족 주식수 숫자들 가져오기
    family_numbers = get_family_share_numbers(excel_path, company)
    # family_numbers가 비어 있으면 라벨링은 불가 → 그냥 테이블만 저장하거나, 건너뛸지 선택
    if not family_numbers:
        print(f"------------- [WARN] 엑셀에서 '{company}'의 친족지분 수식을 찾지 못했습니다. (라벨링 불가)")
        # 필요하다면 여기서 그냥 return 0 해도 됨
        # return 0

    for i, tbl in enumerate(tables, start=1):
        save_this = False
        tid = tbl.get("id", "")
        if tid:
            # id가 있을 때: id가 ...table_002 로 끝나는지 확인
            if re.search(rf"{re.escape(target_suffix)}$", tid):
                save_this = True
        else:
            # id가 없으면 table_2 처럼 인덱스로 판단
            m = re.match(r"table_(\d+)$", target_suffix)
            if m:
                target_idx = int(m.group(1))
                if i == target_idx:
                    save_this = True

        if not save_this:
            continue

        try:
            dfs = pd.read_html(StringIO(str(tbl)), flavor="lxml")
            if not dfs:
                continue

            df = dfs[0]

            # 셀 값 정리 (\r 제거 + strip)
            df = df.apply(
                lambda s: s.map(
                    lambda x: str(x).replace("\r", "").strip()
                    if pd.notnull(x) else x
                )
            )

            # 컬럼 이름 처리 (MultiIndex → 문자열로 flatten)
            if isinstance(df.columns, pd.MultiIndex):
                new_cols = []
                for col in df.columns:
                    # col: ('소유주식수 및 지분율', '기 말', '주식수') 같은 튜플
                    parts = [
                        str(c).strip()
                        for c in col
                        if pd.notnull(c) and str(c).strip() != ""
                    ]
                    new_cols.append("_".join(parts) if parts else "")
                df.columns = new_cols
            else:
                df.columns = df.columns.str.strip()

            # 기본 컬럼명 통일 (성명 / 관계 정도만)
            rename_map = {
                "성 명": "성명",
                "성 명_성 명_성 명": "성명",
                "관 계": "관계",
                "관 계_관 계_관 계": "관계",
                "주식의 종류": "주식의종류",
                "주식의 종류_주식의 종류_주식의 종류": "주식의종류",
            }
            df = df.rename(columns=rename_map)

            # 디버그: 컬럼명 확인
            #print(f"[디버그] {company} 컬럼명: {list(df.columns)}")

            # 🔍 주식수 기준 컬럼 자동 탐색 → '주식수_판단'에 사용
            target_col = None

            # 1순위: '기 말' + '주식수' 둘 다 포함된 컬럼
            candidates = [
                c for c in df.columns
                if ("주식수" in str(c)) and ("기 말" in str(c))
            ]
            if candidates:
                target_col = candidates[0]
            else:
                # 2순위: 이름에 '주식수'만 포함된 컬럼
                candidates = [
                    c for c in df.columns
                    if "주식수" in str(c)
                ]
                if candidates:
                    target_col = candidates[0]

            if target_col is None:
                print(f"[정보] '{company}' 테이블에 주식수 관련 컬럼이 없어 건너뜁니다.")
                continue

            #print(f"[디버그] {company} 주식수 기준 컬럼 선택: {target_col}")

            # 선택된 주식수 컬럼을 숫자로 변환 → '주식수_판단'
            df["주식수_판단"] = pd.to_numeric(
                df[target_col].astype(str).str.replace(",", ""),
                errors="coerce"
            )

            # 성명 컬럼 존재 확인
            if "성명" not in df.columns:
                print(f"[정보] '{company}' 테이블에 '성명' 컬럼이 없어 건너뜁니다.")
                continue

            # 필수 컬럼 결측 제거
            df = df.dropna(subset=["성명", "주식수_판단"])

            # 친족여부 라벨 함수 (허용 오차 포함)
            def label_fn(x):
                if not family_numbers:
                    return None  # 엑셀에서 정보 못 찾은 경우
                try:
                    x_int = int(x)
                except Exception:
                    return None

                # ±(FAMILY_REL_TOL * 100)% 이내면 '친족'
                if is_close_to_any_family_number(x_int, family_numbers, rel_tol=FAMILY_REL_TOL):
                    return "친족"
                else:
                    return "아님"

            # '친족여부' 열 추가
            df["친족여부"] = df["주식수_판단"].apply(label_fn)
            df = df.dropna(subset=["친족여부"])

            # 최종 CSV 저장
            path = os.path.join(out_dir, f"{safe_company}_주주관계.csv")
            df.to_csv(path, index=False, encoding="utf-8-sig")
            print(f"saved: {path}")
            saved += 1

        except Exception as e:
            print(f"table#{i} 저장 실패: {e}")
            continue

    return saved

def extract_and_save_only_target_table_current(
    html_text: str,
    out_dir: str,
    company: str,
    file_prefix: str = "",
    target_suffix: str = "table_002",
) -> int:
    """
    CURRENT 용:
    - HTML에서 target_suffix 테이블만 추출
    - pd.read_html 기본 동작을 그대로 이용해 MultiIndex 헤더를 읽어옴
    - MultiIndex 헤더를 문자열로 flatten
    - '성명', '관계', '주식의종류' 컬럼 이름 정리
    - '주식수_판단' 컬럼 생성 (기말 주식수 기준)
    - 친족여부 라벨링은 하지 않음
    """
    soup = BeautifulSoup(html_text, "lxml")

    # <br> → 줄바꿈
    for br in soup.find_all("br"):
        br.replace_with("\n")

    tables = soup.find_all("table")
    if not tables:
        return 0

    os.makedirs(out_dir, exist_ok=True)
    saved = 0

    # 기업명 안전하게 파일명으로 변환
    safe_company = re.sub(r"[^\w가-힣]", "_", company)

    for i, tbl in enumerate(tables, start=1):
        save_this = False
        tid = tbl.get("id", "")

        if tid:
            # id가 있을 때: id가 ...table_002 로 끝나는지 확인
            if re.search(rf"{re.escape(target_suffix)}$", tid):
                save_this = True
        else:
            # id가 없으면 table_2 처럼 index로 판단
            m = re.match(r"table_(\d+)$", target_suffix)
            if m:
                target_idx = int(m.group(1))
                if i == target_idx:
                    save_this = True

        if not save_this:
            continue

        try:
            # family 버전과 동일하게 기본 read_html 사용
            dfs = pd.read_html(StringIO(str(tbl)), flavor="lxml")
            if not dfs:
                continue

            df = dfs[0]

            # 셀 값 정리 (\r 제거 + strip)
            df = df.apply(
                lambda s: s.map(
                    lambda x: str(x).replace("\r", "").strip()
                    if pd.notnull(x) else x
                )
            )

            # 🔹 MultiIndex 헤더 → 문자열로 flatten
            if isinstance(df.columns, pd.MultiIndex):
                new_cols = []
                for col in df.columns:
                    # col 예: ('소유주식수 및 지분율', '기 말', '주식수')
                    parts = [
                        str(c).strip()
                        for c in col
                        if pd.notnull(c) and str(c).strip() != ""
                    ]
                    new_cols.append("_".join(parts) if parts else "")
                df.columns = new_cols
            else:
                df.columns = df.columns.str.strip()

            # 🔹 기본 컬럼명 통일 (예전 family 함수와 동일한 방식)
            rename_map = {
                "성 명": "성명",
                "성 명_성 명_성 명": "성명",
                "성명_성명_성명": "성명",
                "관 계": "관계",
                "관 계_관 계_관 계": "관계",
                "관계_관계_관계": "관계",
                "주식의 종류": "주식의종류",
                "주식의 종류_주식의 종류_주식의 종류": "주식의종류",
                "주식의종류_주식의종류_주식의종류": "주식의종류",
            }
            df = df.rename(columns=rename_map)

            # 🔍 주식수 기준 컬럼 자동 탐색 (기말 주식수 우선)
            target_col = None

            # 1순위: '기 말' + '주식수' 모두 포함된 컬럼
            candidates = [
                c for c in df.columns
                if ("주식수" in str(c)) and ("기 말" in str(c))
            ]
            if candidates:
                target_col = candidates[0]
            else:
                # 2순위: 이름에 '주식수'만 포함된 컬럼
                candidates = [
                    c for c in df.columns
                    if "주식수" in str(c)
                ]
                if candidates:
                    target_col = candidates[0]

            if target_col is None:
                print(f"[INFO] '{company}' CURRENT 테이블에 주식수 관련 컬럼이 없어 건너뜁니다.")
                continue

            # 선택된 주식수 컬럼을 숫자로 변환 → '주식수_판단'
            df["주식수_판단"] = pd.to_numeric(
                df[target_col].astype(str).str.replace(",", ""),
                errors="coerce"
            )

            # 🔹 혹시라도 헤더 줄이 데이터로 들어온 경우 방어적으로 제거
            # 헤더가 데이터로 들어온 행/숫자 없는 행 정리
            if "성명" in df.columns and "관계" in df.columns:
                df = df[~((df["성명"] == "성명") & (df["관계"] == "관계"))]

            # 🔥 주식수_판단이 NaN인 행(예: '계,계,우선주,-,-,-,-,-') 제거
            df = df[~df["주식수_판단"].isna()]
            df = df.reset_index(drop=True)
            
            # 최종 CSV 저장
            path = os.path.join(out_dir, f"{safe_company}_{file_prefix}{target_suffix}.csv")
            df.to_csv(path, index=False, encoding="utf-8-sig")
            print(f"saved: {path}")
            saved += 1

        except Exception as e:
            print(f"table#{i} CURRENT 저장 실패: {e}")
            continue

    return saved

def main_get_shareholder_current_tables(
    input_data_path: str = INPUT_DATA_PATH,
    out_dir: str = OUTPUT_DIR_PATH_CURRENT,
):
    """
    DART에서 '현재' 주주 테이블을 가져와서 out_dir에 저장하는 버전.
    - 엑셀 기반 '친족여부' 라벨을 붙이지 않고,
      extract_and_save_only_target_table()만 사용해서 순수 테이블만 저장.
    - step2_diff_shareholders.py 에서 CURRENT_DIR를 채우는 용도로 사용.
    """
    df = pd.read_csv(input_data_path)
    print(f"{input_data_path} 내 확인된 총 기업 수: {len(df)}개")

    driver = build_chrome_driver(HEADLESS)
    total_saved = 0

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Create CURRENT Tables"):
        company = str(row.get("회사명", "")).strip()
        dart_url = str(row.get("사업보고서\n바로가기", "")).strip()
        if not dart_url:
            print(f"[{idx+1}/{len(df)}] {company} - DART 링크 없음, 스킵")
            continue

        print(f"\n[{idx+1}/{len(df)}] {company} (CURRENT)")
        print(f"main.do 요청: {dart_url}")

        # ① main.do에서 treeData 블록 파싱
        try:
            r = requests.get(dart_url, timeout=10)
            main_soup = BeautifulSoup(r.text, "html.parser")
            block = find_tree_block_for_target(main_soup, TARGET_TEXT)
            if not block:
                print("[ERROR] 'VII. 주주에 관한 사항' <script> 블록을 찾을 수 없음")
                continue
            params = extract_viewer_params_from_tree_block(block, TARGET_TEXT)
            if not params:
                print("[ERROR] viewer 파라미터 추출 실패 (필수 필드 없음)")
                continue
            viewer_url = build_viewer_url(params)
            print("viewer.do URL 생성:", viewer_url)
        except Exception as e:
            print("------------- [WARN] main.do 요청/파싱 오류:", e)
            continue

        # ② viewer.do 에서 table_002 HTML 가져오기
        try:
            driver.get(normalize_viewer_url(viewer_url))
            time.sleep(WAIT_AFTER_VIEWER)
            html_with_tables = find_html_with_tables_anywhere(driver)

            prefix = f"ele_{int(params['eleId']):03d}_" if params.get("eleId") \
                     else f"idx_{idx+1:03d}_"

            # 수정 후
            saved = extract_and_save_only_target_table_current(
                html_with_tables,
                out_dir=out_dir,
                company=company,
                file_prefix=prefix,
                target_suffix=TARGET_TABLE_SUFFIX,
            )
            print(f"저장된 CURRENT 테이블 수: {saved}")
            total_saved += saved

        except Exception as e:
            print("------------- [WARN] viewer 처리 오류:", e)
            continue

        time.sleep(0.2)

    driver.quit()
    print(f"\nCURRENT 테이블 생성 완료: 총 {total_saved}개 파일 저장 (out_dir={out_dir})")



'''
main 실행 부
'''
def main_get_shareholder_realationship_tables():
    df = pd.read_csv(INPUT_DATA_PATH)
    print(f"{INPUT_DATA_PATH} 내 확인된 총 기업 수: {len(df)}개")

    if os.path.exists(INPUT_EXCEL_PATH):
        print("=" * 50)
        print("    사전 조사된 가족 정보를 활용합니다  ")
        print("=" * 50)
    
    else:
        print("=" * 50)
        print("    사전 조사된 가족 정보를 활용할 수 없습니다  ")
        print("=" * 50)
        
        return

    driver = build_chrome_driver(HEADLESS)
    total_saved = 0

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="Create Relationship Tables"):
        company = str(row.get("회사명", "")).strip()
        dart_url = str(row.get("사업보고서\n바로가기", "")).strip()
        if not dart_url:
            print(f"[{idx+1}/{len(df)}] {company} - DART 링크 없음, 스킵")
            continue

        print(f"\n[{idx+1}/{len(df)}] {company}")
        print(f"main.do 요청: {dart_url}")
        try:
            r = requests.get(dart_url, timeout=10)
            main_soup = BeautifulSoup(r.text, "html.parser")
            block = find_tree_block_for_target(main_soup, TARGET_TEXT)
            if not block:
                print("[ERROR] 'VII. 주주에 관한 사항' <script> 블록을 찾을 수 없음")
                continue
            params = extract_viewer_params_from_tree_block(block, TARGET_TEXT)
            if not params:
                print("[ERROR] viewer 파라미터 추출 실패 (필수 필드 없음)")
                continue
            viewer_url = build_viewer_url(params)
            print("viewer.do URL 생성:", viewer_url)
        except Exception as e:
            print("------------- [WARN] main.do 요청/파싱 오류:", e)
            continue

        try:
            driver.get(normalize_viewer_url(viewer_url))
            time.sleep(WAIT_AFTER_VIEWER)
            html_with_tables = find_html_with_tables_anywhere(driver)
            prefix = f"ele_{int(params['eleId']):03d}_" if params.get("eleId") else f"idx_{idx+1:03d}_"
            saved = extract_and_save_only_target_table_family(
                html_with_tables,
                out_dir=OUTPUT_DIR_PATH,
                company=company,
                excel_path=INPUT_EXCEL_PATH,
                file_prefix=prefix,
                target_suffix=TARGET_TABLE_SUFFIX,
            )
            print(f"저장된 테이블 수: {saved}")
            total_saved += saved

        except Exception as e:
            print("------------- [WARN] viewer 처리 오류:", e)
            continue

        time.sleep(0.2)


    driver.quit()
    print(f"\n전체 완료: 총 {total_saved}개 파일 저장")

if __name__ == "__main__":
    main_get_shareholder_realationship_tables()

